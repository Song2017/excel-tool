import os
from typing import Tuple, List

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.workbook import Workbook
from openpyxl.worksheet.dimensions import SheetFormatProperties
from openpyxl.worksheet.worksheet import Worksheet


def get_excel_range(base: int, height: int, column: str) -> tuple:
    cols = column.split()
    return cols[0] + str(base), cols[1] + str(base + height - 1)


def get_excel_sheet(
        sheet_name: str = "",
        sheet_index: int = 0, file_path='./demo.xlsx') -> Tuple[Workbook, Worksheet]:
    wb = load_workbook(file_path)
    # 选择要追加数据的工作表，这里假设是第一个工作表
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.worksheets[sheet_index]
    return wb, ws


def export_excel(i_data: list, i_columns: list,
                 file_path: str = "./demo.xlsx", sheet_name="order") -> str:
    db = pd.DataFrame(i_data, columns=i_columns, dtype=str)
    mode = "a" if os.path.exists(file_path) else "w"
    ise = "replace" if mode == "a" else None
    with pd.ExcelWriter(file_path, engine="openpyxl", mode=mode, if_sheet_exists=ise) as writer:
        db.to_excel(writer, sheet_name=sheet_name, index=False)
    wb, ws = get_excel_sheet(sheet_name, 0, file_path)
    ws.sheet_format = SheetFormatProperties(defaultColWidth=12, defaultRowHeight=20)
    wb.save(file_path)
    return file_path


def append_table(rows: list, sheet_name: str = "", sheet_index: int = 0, interval: int = 1, file_path='./demo.xlsx'):
    # 加载已存在的工作簿
    wb, ws = get_excel_sheet(sheet_name, sheet_index, file_path)
    # 获取当前工作表中的最大行数
    max_row = ws.max_row
    max_row += interval
    # 追加一行数据，注意行号是从1开始的，所以新行的行号是 max_row + 1
    for col_num, value in enumerate(rows, start=1):
        ws.cell(row=max_row + 1, column=col_num, value=value)
    wb.save(file_path)


def append_table_head(col: int, row: int, values: list, start: str = "A 1", color="709bfa", font_size=12,
                      sheet_name: str = "", sheet_index: int = 0, row_height: int = 30, file_path='./demo.xlsx'):
    wb, ws = get_excel_sheet(sheet_name, sheet_index, file_path)
    for i in range(len(values)):
        cell = ws.cell(row, col + i)
        cell.value = values[i]
        cell.font = Font(size=font_size, bold=True)
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    ws.row_dimensions[row].height = row_height
    wb.save(file_path)


def append_sheet_title(start: str, end: str, value: str = "", sheet_name: str = "", sheet_index: int = 0,
                       file_path='./demo.xlsx', color="709bfa", font_size=12):
    # 创建一个新的工作簿
    # 加载已存在的工作簿
    wb, ws = get_excel_sheet(sheet_name, sheet_index, file_path)

    # 设置填充颜色为蓝色
    fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    # 合并单元格，从 A1 到 C2（跨3列，高2行）
    ws.merge_cells(f'{start}:{end}')
    # 选择合并后的单元格（合并后仍然可以通过左上角的单元格引用）
    merged_cell = ws[start]
    # 设置单元格填充颜色
    merged_cell.fill = fill
    # 设置文本内容
    merged_cell.value = value or '文本'
    merged_cell.font = Font(size=font_size, bold=True)
    # 设置文本内容居中
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    # 保存工作簿
    wb.save(file_path)


def write_cells(col: int, row: int, data: List,
                color="709bfa", font_size=12,
                sheet_name: str = "", sheet_index: int = 0, row_height: int = 80, file_path='./demo.xlsx'):
    assert len(data) > 0, "write_cells: please check data"
    row_len = len(data)
    col_len = len(data[0])
    wb, ws = get_excel_sheet(sheet_name, sheet_index, file_path)
    for r in range(row_len):
        for c in range(col_len):
            cell = ws.cell(row + r, col + c)
            cell.value = data[r][c]
            cell.font = Font(size=font_size)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    wb.save(file_path)


if __name__ == '__main__':
    export_excel([1, 2, 3], ['column'])
    export_excel([1, 2, 3], ['column'], sheet_name="ss")
    # append_table([1, 2, 3])
    append_sheet_title("A1", "D2")
